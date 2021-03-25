import csv
import datetime
import os
import re
import shutil
import tempfile
import unicodedata
import urllib.request
from typing import Any, Optional

from openpyxl import load_workbook


def fetch_latest_cities_file_url():
    with urllib.request.urlopen(
        "https://www.soumu.go.jp/denshijiti/code.html"
    ) as res:
        body = res.read()

    # 最新のファイル名を特定
    m = re.search(b"/main_content/\\d{9}.xlsx", body)
    if not m:
        raise Exception
    return f"https://www.soumu.go.jp{m.group().decode()}"


def fetch_latest_unions_file_url():
    with urllib.request.urlopen(
        "https://www.soumu.go.jp/denshijiti/code.html"
    ) as res:
        body = res.read()

    # 最新のファイル名を特定
    m = re.findall(b"/main_content/\\d{9}.xlsx", body)[-1]

    return f"https://www.soumu.go.jp{m.decode()}"


def download_file(url: str, file: str):
    with urllib.request.urlopen(url) as source, open(
        file, "wb"
    ) as destination:
        shutil.copyfileobj(source, destination)


def extract_csvs_from_cities_excel(xlsx_file: str, destination: str):
    workbook = load_workbook(filename=xlsx_file, read_only=True)
    for worksheet, file_name in zip(
        workbook.worksheets, ("designated_cities.csv", "local_governments.csv")
    ):

        rows = worksheet.rows
        # remove headers
        next(rows)
        with open(os.path.join(destination, file_name), "w") as f:
            c = csv.writer(f)
            c.writerow(
                (
                    "code",
                    "prefecture",
                    "city",
                    "prefecture__kana",
                    "city__kana",
                )
            )
            c.writerows(
                (
                    row[0].value,
                    row[1].value,
                    row[2].value,
                    unicodedata.normalize("NFKC", row[3].value)
                    if row[3].value
                    else None,
                    unicodedata.normalize("NFKC", row[4].value)
                    if row[4].value
                    else None,
                )
                for row in rows
                if any(cell.value for cell in row)
            )


def _parse_establishment_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    # 明治5年12月2日（1872年12月31日）まで旧暦
    if value == "明治5年4月16日":
        return datetime.date(1872, 5, 22)
    value_type = type(value)
    if value_type == datetime.datetime:
        return value.date()
    if value_type == int:
        return datetime.date(1899, 12, 30) + datetime.timedelta(value)
    if value_type != str:
        raise NotImplementedError
    if value.startswith("明治"):
        m = re.search("明治(\d{1,2})年(\d{1,2})月(\d{1,2})日", value)
        if int(m.group(1)) <= 5:
            raise NotImplementedError
        return datetime.date(
            1867 + int(m.group(1)), int(m.group(2)), int(m.group(3))
        )
    else:
        raise NotImplementedError


def extract_csv_from_unions_excel(xlsx_file: str, destination: str):
    workbook = load_workbook(filename=xlsx_file, read_only=True)
    worksheet = workbook.worksheets[0]

    rows = worksheet.rows
    # remove headers
    for _ in range(10):
        next(rows)

    with open(os.path.join(destination, "unions.csv"), "w") as f:
        c = csv.writer(f)
        c.writerow(
            (
                "code",
                "name",
                "name__kana",
                "establishment_date",
                "zipcode",
                "address",
                "telephone_number",
            )
        )
        c.writerows(
            (
                row[1].value,
                row[2].value,
                row[3].value,
                _parse_establishment_date(row[4].value),
                row[5].value,
                row[6].value,
                row[7].value,
            )
            for row in rows
        )


def copy_schema_file(destination: str, file: str):
    shutil.copyfile(
        os.path.join(
            os.path.dirname(__file__),
            file,
        ),
        os.path.join(destination, file),
    )


def main():
    destination = os.environ["DESTINATION"]
    cities_file_url = fetch_latest_cities_file_url()
    unions_file_url = fetch_latest_unions_file_url()
    os.makedirs(destination, exist_ok=True)
    with tempfile.TemporaryDirectory() as tmpdirname:
        download_file(
            cities_file_url,
            os.path.join(tmpdirname, "city_codes.xlsx"),
        )
        download_file(
            unions_file_url,
            os.path.join(tmpdirname, "union_codes.xlsx"),
        )
        extract_csvs_from_cities_excel(
            os.path.join(tmpdirname, "city_codes.xlsx"), destination
        )
        extract_csv_from_unions_excel(
            os.path.join(tmpdirname, "union_codes.xlsx"), destination
        )
    copy_schema_file(destination, "designated_cities_schema.json")
    copy_schema_file(destination, "local_governments_schema.json")
    copy_schema_file(destination, "unions_schema.json")


if __name__ == "__main__":
    main()
